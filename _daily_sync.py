# -*- coding: utf-8 -*-
"""
_daily_sync.py — 매일 07:45 자동 실행
전날 생산수불 Firestore → 보고서 원본 자동 추가 + 텔레그램 알림
실행: py _daily_sync.py              (어제 날짜)
     py _daily_sync.py 2026-06-19   (날짜 지정 테스트)
"""
import sys, os, json, re, datetime, time, urllib.request, urllib.error
from copy import copy
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# ── 설정 ────────────────────────────────────────────────────
TG_TOKEN   = "8863160843:AAGhaaRny1Uu225mM1C9vYZiJKa2awixd7s"
CHAT_IDS   = ["7254187932"]

API_KEY    = "AIzaSyBoxMBCwOqqmjpOAHVKmY5twJeZSanE_6E"
FS_BASE    = "https://firestore.googleapis.com/v1/projects/vicon-3factory/databases/(default)/documents"

WO_FOLDERS = [
    r"C:\Users\user\Desktop\비콘공용폴더\1. 생산관리-일체형\1. 작업지시서",
    r"C:\Users\user\Desktop\비콘공용폴더\4. 생산관리-탈형",
]
REPORT_DIR = r"C:\Users\user\Desktop\비콘공용폴더"
DATA_START = 4
RED_FILL   = PatternFill("solid", fgColor="FFFF0000")

# ── 날짜 결정 ────────────────────────────────────────────────
if len(sys.argv) > 1:
    TARGET = sys.argv[1]
    target_date = datetime.datetime.strptime(TARGET, "%Y-%m-%d").date()
else:
    target_date = datetime.date.today() - datetime.timedelta(days=1)
    TARGET = target_date.strftime("%Y-%m-%d")

t0 = time.time()
def tlog(msg): print(f"[+{time.time()-t0:5.1f}s] {msg}", flush=True)
tlog(f"대상 날짜: {TARGET}")

# ── 텔레그램 ─────────────────────────────────────────────────
def tg(text):
    for cid in CHAT_IDS:
        try:
            d = json.dumps({"chat_id": cid, "text": text}).encode()
            req = urllib.request.Request(
                f"https://api.telegram.org/bot{TG_TOKEN}/sendMessage",
                data=d, method="POST")
            req.add_header("Content-Type", "application/json")
            urllib.request.urlopen(req, timeout=10)
        except Exception as e:
            print(f"TG 오류: {e}")

# ── Firestore 조회 ───────────────────────────────────────────
def fetch(date_str):
    q = {"structuredQuery": {
        "from": [{"collectionId": "production_deck"}],
        "where": {"fieldFilter": {"field": {"fieldPath": "date"},
                  "op": "EQUAL", "value": {"stringValue": date_str}}},
        "orderBy": [{"field": {"fieldPath": "date"}, "direction": "ASCENDING"}]
    }}
    d = json.dumps(q).encode()
    req = urllib.request.Request(FS_BASE + ":runQuery?key=" + API_KEY, data=d, method="POST")
    req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            body = json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"Firestore {e.code}: {e.read().decode()[:200]}")
    recs = []
    for item in body:
        doc = item.get("document")
        if not doc: continue
        f = doc.get("fields", {})
        def s(k): v = f.get(k, {}); return v.get("stringValue", "") or str(v.get("integerValue", ""))
        pv = f.get("packs", {}).get("arrayValue", {}).get("values", [])
        packs = []; pack_details = {}
        for p in pv:
            mp = p.get("mapValue", {}).get("fields", {}); vv = mp.get("pack_no", {})
            pno = None
            if "integerValue" in vv: pno = int(vv["integerValue"])
            elif "doubleValue" in vv: pno = int(vv["doubleValue"])
            if pno is not None:
                packs.append(pno)
                sv = mp.get("sheets", {}); av = mp.get("area", {})
                sheets = int(float(sv.get("integerValue") or sv.get("doubleValue") or 0))
                area   = round(float(av.get("integerValue") or av.get("doubleValue") or 0), 3)
                pack_details[pno] = {"sheets": sheets, "area": area}
        recs.append({"id": doc["name"].split("/")[-1],
                     "date": s("date"), "machine": s("machine"), "shift": s("shift"),
                     "company": s("company"), "site": s("site"), "req_code": s("req_code"),
                     "recorded_by": s("recorded_by"),
                     "packs": sorted([p for p in packs if p is not None]),
                     "pack_details": pack_details})
    return recs

# ── 보고서 파일 탐색 ─────────────────────────────────────────
def find_report(month):
    cands = []
    for fn in os.listdir(REPORT_DIR):
        m = re.match(r"3공장업무보고서\((\d+)월\)\.(xlsx|xlsm)$", fn)
        if m:
            cands.append((int(m.group(1)), os.path.join(REPORT_DIR, fn)))
    if not cands: return None
    # 해당 월 우선, 없으면 가장 최신 월
    same = [p for mo, p in cands if mo == month]
    return same[0] if same else sorted(cands)[-1][1]

# ── 작업지시서 인덱스 + 매칭 ─────────────────────────────────
def wo_index():
    idx = []
    for base in WO_FOLDERS:
        for root, _, files in os.walk(base):
            for fn in files:
                if not fn.lower().endswith((".xlsm", ".xlsx", ".xls")): continue
                full = os.path.join(root, fn)
                try: mt = os.path.getmtime(full)
                except: mt = 0
                idx.append({"fn": fn, "full": full, "mtime": mt, "old": "구버전" in root})
    return idx

def find_wo(req, packs, idx):
    cands = []
    for it in idx:
        nm = it["fn"]; i = nm.find(req)
        while i != -1:
            a = nm[i+len(req):i+len(req)+1]
            if a == "" or a in ", ()_.([": cands.append(it); break
            i = nm.find(req, i+1)
    if not cands: return None
    pm, px = (min(packs), max(packs)) if packs else (None, None)
    for c in cands:
        m = re.match(r'\s*(\d+)\s*(?:[-~]\s*(\d+))?\s*[#@,]', c["fn"])
        rng = (int(m.group(1)), int(m.group(2)) if m.group(2) else int(m.group(1))) if m else None
        c["_ok"] = bool(rng and pm and rng[0] <= pm and px <= rng[1])
    cands.sort(key=lambda x: (not x["_ok"], x["old"], -x["mtime"]))
    return cands[0]

# ── WO 로드 ──────────────────────────────────────────────────
def gv(rd, col):
    info = rd.get(col)
    if info is None: return None
    v = info["v"] if isinstance(info, dict) else info
    return None if (isinstance(v, str) and not v.strip()) else v

def is_top_f(rd, col):
    info = rd.get(col); return info.get("is_top", True) if info else False

def load_wo(path):
    wb = load_workbook(path, data_only=True)
    ws = next((wb[nm] for nm in wb.sheetnames if nm.strip().startswith("DECK")), wb[wb.sheetnames[0]])
    title = str(ws.cell(row=1, column=1).value or "").upper()
    wt = "giga" if "GIGA" in title else ("mega" if "MEGA" in title else None)
    merged = {}
    for mr in ws.merged_cells.ranges:
        c0, r0, c1, r1 = mr.bounds; tl = ws.cell(row=r0, column=c0)
        for r in range(r0, r1+1):
            for c in range(c0, c1+1):
                merged[(r, c)] = {"v": tl.value, "is_top": (r == r0 and c == c0)}
    grid = {}
    for row in ws.iter_rows(min_row=1, max_row=500, max_col=20):
        for cell in row:
            key = (cell.row, cell.column); cl = get_column_letter(cell.column)
            if key in merged:
                info = merged[key]
                if info["v"] is not None or not info["is_top"]:
                    grid.setdefault(cell.row, {})[cl] = info
            elif cell.value is not None:
                grid.setdefault(cell.row, {})[cl] = {"v": cell.value, "is_top": True}
    wb.close()
    if wt is None:
        bc = sum(1 for r in grid if isinstance(gv(grid[r], "B"), (int, float)))
        cc = sum(1 for r in grid if isinstance(gv(grid[r], "C"), (int, float)))
        wt = "giga" if cc > bc else "mega"
    return grid, wt

def extract(grid, is_giga, packs):
    pc = "C" if is_giga else "B"; out = []; last = None
    for r in sorted(grid):
        if r < 9: continue
        rd = grid[r]; a = gv(rd, "A")
        if a is None and gv(rd, pc) is None: continue
        if isinstance(a, str) and a.replace(" ", "").startswith("합계"): continue
        pk = gv(rd, pc); top = is_top_f(rd, pc) if pk is not None else False
        if pk is not None and top:
            try: last = int(pk)
            except: pass
        if last in packs and a is not None:
            out.append((rd, last if (top and pk is not None) else None))
    return out

# ── 수량·면적 불일치 체크 ────────────────────────────────────
def check_qty(grid, is_giga, pack_details):
    """작업지시서 TG별 수량·면적(TG=0 제외) vs 앱 등록값 비교"""
    pack_col = "C" if is_giga else "B"
    issues = []
    cur_pack = None; wo_qty = 0; wo_area = 0.0

    def flush():
        nonlocal cur_pack, wo_qty, wo_area
        if cur_pack is not None and cur_pack in pack_details:
            app = pack_details[cur_pack]; diffs = []
            if wo_qty != app["sheets"]:
                diffs.append(f"수량 WO:{wo_qty}매 ≠ 앱:{app['sheets']}매")
            if abs(wo_area - app["area"]) > 0.1:
                diffs.append(f"면적 WO:{round(wo_area,3)} ≠ 앱:{app['area']}")
            if diffs:
                issues.append(f"패킹{cur_pack} / " + " / ".join(diffs))
        cur_pack = None; wo_qty = 0; wo_area = 0.0

    for r in sorted(grid.keys()):
        if r < 9: continue
        rd = grid[r]
        pk = gv(rd, pack_col)
        if pk is not None and is_top_f(rd, pack_col):
            flush()
            try: cur_pack = int(pk)
            except: cur_pack = None
        if cur_pack not in pack_details: continue
        try:
            tg = gv(rd, "H")
            if tg is not None and int(float(str(tg))) == 0: continue  # TG=0 제외
        except: pass
        try:
            q = gv(rd, "J")
            if q is not None: wo_qty += int(float(str(q)))
        except: pass
        try:
            a = gv(rd, "K")
            if a is not None: wo_area += float(str(a))
        except: pass
    flush()
    return issues

# ── 기존 중복 스캔 ───────────────────────────────────────────
def scan_exist(ws, is_giga, date_str):
    # key: (req_code, machine, shift)
    # 메가: D=4, Q=17(호기), R=18(주야) / 기가: D=4, S=19(주야)
    req_c = 4
    m_c   = None if is_giga else 17
    s_c   = 19 if is_giga else 18
    keys = set()
    cap = min(ws.max_row or 60000, 60000)
    for row in ws.iter_rows(min_row=DATA_START, max_row=cap,
                             min_col=1, max_col=20, values_only=True):
        d = row[0]
        if d is None: continue
        ds = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)[:10]
        if ds != date_str: continue
        req = row[req_c-1]
        machine = str(row[m_c-1]) if m_c and row[m_c-1] else ""
        shift   = str(row[s_c-1]) if row[s_c-1] else ""
        if req: keys.add((str(req), machine, shift))
    return keys

# ── 서식 캡처 + 마지막행 ─────────────────────────────────────
def cap_styles(ws, ncol, ref=9):
    st = {}
    for c in range(1, ncol+1):
        cell = ws.cell(row=ref, column=c)
        st[c] = {"font": copy(cell.font), "fill": copy(cell.fill),
                 "nf": cell.number_format,
                 "h": cell.alignment.horizontal, "v": cell.alignment.vertical or "center",
                 "shrink": cell.alignment.shrink_to_fit or False,
                 "wrap": cell.alignment.wrap_text or False}
    return st

def last_data_row(ws, key_col=4):
    last = DATA_START - 1; cap = min(ws.max_row or 60000, 60000)
    for i, row in enumerate(ws.iter_rows(min_row=DATA_START, max_row=cap,
                                          min_col=key_col, max_col=key_col, values_only=True)):
        if row[0] not in (None, ""): last = DATA_START + i
    return last

def mno(m): n = re.sub(r'[^0-9]', '', str(m)); return int(n) if n else 0

# ── 시트에 행 추가 ───────────────────────────────────────────
def append_block(ws, segs, start_r, is_giga, styles, is_mismatch, rec):
    pack_col  = 7 if is_giga else 6   # G 또는 F
    struct_r  = 18 if is_giga else 17  # 테두리 실선 시작 (우측)
    t_col     = 20
    d = datetime.datetime.strptime(rec["date"], "%Y-%m-%d").date()
    mn = mno(rec["machine"])

    # 행 값 목록
    rows_vals = []
    for rd, pm in segs:
        showpk = pm is not None
        if is_giga:
            v = [d, rec["company"], rec["site"], rec["req_code"],
                 gv(rd,"A"), gv(rd,"B"), gv(rd,"C") if showpk else None,
                 gv(rd,"D"), gv(rd,"E"), gv(rd,"F"), gv(rd,"G"),
                 gv(rd,"H"), gv(rd,"I"), gv(rd,"J"), gv(rd,"K"),
                 gv(rd,"M"), gv(rd,"N"), "티아이", rec["shift"], None]
        else:
            v = [d, rec["company"], rec["site"], rec["req_code"],
                 gv(rd,"A"), gv(rd,"B") if showpk else None, gv(rd,"C"),
                 gv(rd,"D"), gv(rd,"E"), gv(rd,"F"), gv(rd,"G"),
                 gv(rd,"H"), gv(rd,"I"), gv(rd,"J"),
                 gv(rd,"L"), gv(rd,"M"), mn, rec["shift"], "티아이", None]
        rows_vals.append((v, pm))

    # 팩 그룹 경계
    grp = {}  # row_idx -> (is_first, is_last)
    merges = []
    i = 0
    while i < len(segs):
        _, pm = segs[i]
        if pm is not None:
            span = 1; j = i + 1
            while j < len(segs) and segs[j][1] is None: span += 1; j += 1
            if span > 1: merges.append((i, span))
            for k in range(span): grp[i+k] = (k == 0, k == span-1)
            i = j
        else:
            i += 1

    # 셀 쓰기
    for ri, (vals, _) in enumerate(rows_vals):
        r = start_r + ri
        is_first, is_last = grp.get(ri, (True, True))
        in_grp = ri in grp

        # TG=0 여부: 메가=vals[10](K=TG개수), 기가=vals[11](L=TG개수)
        tg_idx = 11 if is_giga else 10
        try:    is_tg0 = int(float(str(vals[tg_idx]))) == 0
        except: is_tg0 = False

        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci)
            if ci == t_col:
                cell.value = f"=+L{r}*M{r}*N{r}*0.2/1000" if is_giga else f"=+K{r}*L{r}*M{r}*0.2/1000"
            else:
                cell.value = val
            st = styles[ci]
            # Fill: mismatch 전행 빨강 / TG=0 K~M(11~13) 빨강 / 그 외 원본
            if is_mismatch or (is_tg0 and 11 <= ci <= 13):
                cell.fill = RED_FILL
            else:
                cell.fill = copy(st["fill"])
            # Font: TG=0 K~M 흰글자 굵게 (mismatch 행 제외)
            if is_tg0 and 11 <= ci <= 13 and not is_mismatch:
                orig = st["font"]
                cell.font = Font(bold=True, color="FFFFFFFF",
                                 name=orig.name, size=orig.size)
            else:
                cell.font = copy(st["font"])
            cell.number_format = st["nf"]
            # Alignment: 패킹열 가운데 / shrink_to_fit·wrap 원본 반영
            h = "center" if ci == pack_col else st["h"]
            cell.alignment = Alignment(horizontal=h, vertical=st["v"] or "center",
                                       shrink_to_fit=st["shrink"], wrap_text=st["wrap"])
            in_solid = (ci <= 4) or (ci >= struct_r)
            top_s = "thin" if (in_solid or is_first or not in_grp) else "dotted"
            bot_s = "thin" if (in_solid or is_last  or not in_grp) else "dotted"
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style=top_s), bottom=Side(style=bot_s))
        ws.cell(row=r, column=1).number_format = "YYYY-MM-DD"

    # 팩열 병합
    for mi, span in merges:
        r1 = start_r + mi; r2 = r1 + span - 1
        ws.merge_cells(start_row=r1, start_column=pack_col, end_row=r2, end_column=pack_col)

    return len(rows_vals)

# ════════════════════════════════════════════════════════════
# 메인
# ════════════════════════════════════════════════════════════

# 1) Firestore 조회
tlog("Firestore 조회 중...")
try:
    recs = fetch(TARGET)
except Exception as e:
    msg = f"[비콘 3공장] ❌ {TARGET} 동기화 실패\nFirestore 오류: {e}"
    tg(msg); sys.exit(1)
tlog(f"  {len(recs)}건")

if not recs:
    tg(f"[비콘 3공장] {TARGET}\n생산 데이터 없음 - 패스")
    sys.exit(0)

# 2) WO 인덱스 + 매칭
tlog("작업지시서 인덱스 구성...")
idx = wo_index()
tlog(f"  {len(idx)}개 파일")
matched = []
for rec in recs:
    wo = find_wo(rec["req_code"], rec["packs"], idx)
    matched.append({**rec, "wo_file": wo["full"] if wo else None})

# 3) WO 병렬 로드
paths = list({r["wo_file"] for r in matched if r.get("wo_file")})
tlog(f"WO {len(paths)}개 로드 중...")
_wc = {}; _wt = {}
with ThreadPoolExecutor(max_workers=8) as ex:
    for p, (g, t2) in zip(paths, ex.map(load_wo, paths)):
        _wc[p] = g; _wt[p] = t2
tlog("WO 로드 완료")

# 4) 보고서 파일
report = find_report(target_date.month)
if not report:
    tg(f"[비콘 3공장] ❌ 보고서 파일 없음 ({target_date.month}월)")
    sys.exit(1)
tlog(f"보고서: {os.path.basename(report)}")

wb = load_workbook(report)
ws_M = wb["메가데크생산수불현황"]
ws_G = wb["기가데크생산수불현황"]

# 5) 중복 스캔
exist_M = scan_exist(ws_M, False, TARGET)
exist_G = scan_exist(ws_G, True,  TARGET)
tlog(f"기존 중복키: 메가 {len(exist_M)} / 기가 {len(exist_G)}")

# 6) 서식 + 마지막행
st_M = cap_styles(ws_M, 20)
st_G = cap_styles(ws_G, 20)
lr_M = last_data_row(ws_M)
lr_G = last_data_row(ws_G)
tlog(f"마지막행: 메가 {lr_M} / 기가 {lr_G}")

# 7) 정렬 + 추가
SHIFT_ORD = {"주간": 0, "야간": 1}
recs_s = sorted(matched, key=lambda x: (x["date"], mno(x["machine"]), SHIFT_ORD.get(x["shift"], 9)))

add_M = add_G = skip = 0
mismatches = []
skip_list = []
qty_issues = []   # 수량·면적 불일치

# req_code+wo_file 단위로 pack_details 합산 (주간+야간 합쳐서 WO 전체와 비교)
from collections import defaultdict
_agg_pd = defaultdict(lambda: defaultdict(lambda: {"sheets": 0, "area": 0.0}))
for _r in matched:
    _k = (_r["req_code"], _r.get("wo_file"))
    for _pno, _pd in (_r.get("pack_details") or {}).items():
        _agg_pd[_k][_pno]["sheets"] += _pd["sheets"]
        _agg_pd[_k][_pno]["area"]    = round(_agg_pd[_k][_pno]["area"] + _pd["area"], 3)
_qty_checked = set()  # 이미 체크한 (req_code, wo_file)

for rec in recs_s:
    wo_file  = rec.get("wo_file")
    is_giga  = (_wt.get(wo_file) == "giga") if wo_file else (rec["machine"] == "탈형")
    ws       = ws_G if is_giga else ws_M
    st       = st_G if is_giga else st_M
    exist    = exist_G if is_giga else exist_M

    mn_str   = str(mno(rec["machine"])) if not is_giga else ""
    dup_key  = (rec["req_code"], mn_str, rec["shift"])

    # 중복 → 스킵
    if dup_key in exist:
        skip += 1
        skip_list.append(f"{rec['machine']} {rec['shift']} / {rec['req_code']} 패킹{rec['packs']}")
        tlog(f"  스킵(중복): {rec['req_code']} {rec['machine']} {rec['shift']}")
        continue

    mismatch = not wo_file
    if mismatch:
        mismatches.append(f"{rec['machine']} {rec['req_code']} 패킹{rec['packs']}")
        # WO 없으면 기본 1행 (빨간 배경)
        d = datetime.datetime.strptime(rec["date"], "%Y-%m-%d").date()
        if is_giga:
            vals = [d, rec["company"], rec["site"], rec["req_code"],
                    None,None,None,None,None,None,None,None,None,None,None,None,None,
                    "티아이", rec["shift"], None]
        else:
            vals = [d, rec["company"], rec["site"], rec["req_code"],
                    None,None,None,None,None,None,None,None,None,None,None,None,
                    mno(rec["machine"]), rec["shift"], "티아이", None]
        segs = [(dict(zip("ABCDEFGHIJKLMNOPQRST",vals)),None)]  # dummy
        # 간단히 직접 씀
        r = (lr_G if is_giga else lr_M) + 1
        struct_r = 18 if is_giga else 17
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=ci)
            if ci == 20:
                cell.value = f"=+L{r}*M{r}*N{r}*0.2/1000" if is_giga else f"=+K{r}*L{r}*M{r}*0.2/1000"
            else:
                cell.value = val
            sv = st[ci]
            cell.font = copy(sv["font"]); cell.fill = RED_FILL
            cell.number_format = sv["nf"]
            cell.alignment = Alignment(horizontal=sv["h"], vertical=sv["v"] or "center")
            cell.border = Border(left=Side(style="thin"),right=Side(style="thin"),
                                 top=Side(style="thin"),bottom=Side(style="thin"))
        ws.cell(row=r, column=1).number_format = "YYYY-MM-DD"
        n = 1
    else:
        grid = _wc[wo_file]
        # 수량·면적 불일치 체크 — 주간+야간 합산 후 WO와 1회만 비교
        _agg_key = (rec["req_code"], wo_file)
        if _agg_key not in _qty_checked:
            _qty_checked.add(_agg_key)
            issues = check_qty(grid, is_giga, dict(_agg_pd[_agg_key]))
            for iss in issues:
                qty_issues.append(f"{rec['req_code']} / {iss}")
        segs = extract(grid, is_giga, set(rec["packs"]))
        if not segs:
            tlog(f"  세그 없음: {rec['req_code']} 패킹{rec['packs']}")
            continue
        start = (lr_G if is_giga else lr_M) + 1
        n = append_block(ws, segs, start, is_giga, st, False, rec)

    if is_giga: lr_G += n; add_G += n
    else:        lr_M += n; add_M += n
    exist.add(dup_key)

# 8) 저장 (파일 잠금 시 1분 간격 최대 5회 재시도)
tlog("저장 중...")
saved = False
for attempt in range(1, 6):
    try:
        wb.save(report)
        saved = True
        tlog(f"저장 완료 (시도 {attempt}회)")
        break
    except PermissionError:
        if attempt < 5:
            tlog(f"파일 잠김 - {attempt}회 시도 실패, 1분 후 재시도...")
            time.sleep(60)
        else:
            msg = (f"[비콘 3공장] {TARGET} 업무보고서 저장 실패\n"
                   f"업무보고서 파일이 열려있어 5회 시도 후 최종 실패.\n"
                   f"파일을 닫고 수동으로 다시 실행해주세요.\n"
                   f"파일: {os.path.basename(report)}")
            tg(msg)
            raise SystemExit(msg)
if not saved:
    raise SystemExit("저장 실패")

# 9) 텔레그램 알림
if add_M == 0 and add_G == 0 and not mismatches:
    lines = [f"[비콘 3공장] {TARGET}", f"전부 중복 - 추가 없음 (스킵 {skip}건)"]
    if skip_list:
        lines.append("▶ 스킵 목록:")
        for s in skip_list: lines.append(f"  - {s}")
else:
    lines = [f"[비콘 3공장] {TARGET} 동기화 완료",
             f"메가데크 +{add_M}행  /  기가데크 +{add_G}행"]
    if skip:
        lines.append(f"중복 스킵: {skip}건")
        for s in skip_list: lines.append(f"  - {s}")
    if mismatches:
        lines.append(f"작업지시서 불일치 {len(mismatches)}건 (빨간행):")
        show_m = mismatches[:10]
        for m in show_m: lines.append(f"  - {m}")
        if len(mismatches) > 10:
            lines.append(f"  ... 외 {len(mismatches)-10}건 생략")
    if qty_issues:
        lines.append(f"수량·면적 불일치 {len(qty_issues)}건:")
        show_q = qty_issues[:10]
        for q in show_q: lines.append(f"  - {q}")
        if len(qty_issues) > 10:
            lines.append(f"  ... 외 {len(qty_issues)-10}건 생략")

tg("\n".join(lines))
tlog("완료: " + " / ".join(lines[1:]))
