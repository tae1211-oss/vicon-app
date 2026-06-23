# -*- coding: utf-8 -*-
"""
작업지시서(면적 포함) → Firestore `job_orders` 업로더  [PC 전용]

흐름 (②안: 생산계획표 기준):
  1) 생산계획표(.xlsm)의 '갑지' 시트들에서 요청코드(req_code) 목록 수집
  2) 네트워크 작업지시서 폴더를 스캔해 req_code 가 파일명에 포함된 최신 파일 찾기
  3) 그 작업지시서의 'DECK작업지시서' 시트에서
     C2=업체 C3=현장 C4=요청코드 / B9~=패킹번호 / N=장수 O=면적  읽어
     QR 과 동일한 d 배열([[장수,면적],...]) 구성
  4) Firestore job_orders 컬렉션에 업로드 (문서ID = makeDocId(업체,현장), 앱과 동일)

앱(index.html)의 생산입력 화면에서 이 목록을 탭하면 QR 스캔 없이 데크 생산입력 폼이 열림.

────────────────────────────────────────────────────────────
사전 준비 (최초 1회):
  1) pip install firebase-admin openpyxl
  2) Firebase 콘솔 → 프로젝트 설정 → 서비스 계정 → "새 비공개 키 생성"
     → 받은 JSON 파일 경로를 아래 SERVICE_ACCOUNT_KEY 에 지정
실행:
  py _upload_job_orders.py            # 실제 업로드
  py _upload_job_orders.py --dry-run  # 업로드 없이 추출 결과만 출력(검증용)
  py _upload_job_orders.py --file "경로\1-6#....xlsm"  # 파일 1개 추출 테스트
────────────────────────────────────────────────────────────
"""
import os, re, sys, glob, datetime, json, urllib.request, urllib.error

# ── 설정 ─────────────────────────────────────────────────────
SERVICE_ACCOUNT_KEY = r"C:\Users\user\Desktop\클로드\vicon-service-account.json"

def _find_plan_file():
    """생산계획표 폴더에서 가장 최근 수정된 데크 생산계획표 파일을 자동 탐지"""
    plan_dir = r"C:\Users\user\Desktop\비콘공용폴더\1. 생산관리-일체형\2. 생산계획표"
    candidates = []
    try:
        for f in os.scandir(plan_dir):
            nm = f.name
            if nm.startswith("~$"): continue
            if "데크 생산계획표" in nm and nm.endswith((".xlsm", ".xlsx", ".xls")):
                candidates.append((f.path, f.stat().st_mtime))
    except OSError:
        pass
    if not candidates:
        raise FileNotFoundError("데크 생산계획표 파일을 찾을 수 없습니다: " + plan_dir)
    return max(candidates, key=lambda x: x[1])[0]

PLAN_FILE = _find_plan_file()

TG_TOKEN = "8863160843:AAGhaaRny1Uu225mM1C9vYZiJKa2awixd7s"
CHAT_IDS = ["7254187932"]

def tg(text):
    for cid in CHAT_IDS:
        try:
            d = json.dumps({"chat_id": cid, "text": text}).encode()
            req = urllib.request.Request(
                f"https://api.telegram.org/bot{TG_TOKEN}/sendMessage",
                data=d, method="POST")
            req.add_header("Content-Type", "application/json")
            urllib.request.urlopen(req, timeout=10)
        except Exception:
            pass
WORK_ORDER_DIRS     = [
    r"C:\Users\user\Desktop\비콘공용폴더",
    r"C:\Users\user\Desktop\비콘공용폴더\4. 생산관리-탈형\1. 작업지시서",
]
SEARCH_DEPTH        = 5          # 폴더 재귀 최대 깊이
PROJECT_ID          = "vicon-3factory"

WO_SHEET   = "DECK작업지시서"
COL_PACK   = 2     # B = 패킹번호
COL_SHEETS = 14    # N = 수량(장수)
COL_AREA   = 15    # O = 면적
ROW_START  = 9


def make_doc_id(company, site):
    """앱 makeDocId 와 동일: 업체|현장, 특수문자 치환, 200자 제한"""
    s = (company or "") + "|" + (site or "")
    s = re.sub(r'[\\/.#\[\]*?]', '_', s)
    return s[:200]


# 갑지 시트 → 호기 그룹 매핑 (데크#1·#2=1,2호기 / 데크#3=3,4호기 / 탈형=탈형)
def machine_group_for_sheet(name):
    if "탈형" in name:
        return "탈형"
    m = re.search(r"갑지\s*(\d+)", name)
    if m:
        return "1,2호기" if int(m.group(1)) <= 2 else "3,4호기"
    return ""


# ── 1) 생산계획표 갑지/탈형 시트에서 요청코드 + 호기그룹 수집 ──
def collect_req_codes(plan_path):
    import openpyxl
    wb = openpyxl.load_workbook(plan_path, data_only=True, read_only=True)
    result = {}   # req_code -> machine_group (먼저 나온 시트 우선)
    for nm in wb.sheetnames:
        group = machine_group_for_sheet(nm)
        if not group:
            continue
        ws = wb[nm]
        # 헤더행에서 '요청코드' 열 찾기
        req_col = None
        header_row = None
        for r, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
            for ci, v in enumerate(row, 1):
                if isinstance(v, str) and v.replace(" ", "") in ("요청코드", "의뢰번호"):
                    req_col, header_row = ci, r
                    break
            if req_col:
                break
        if not req_col:
            continue
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row or header_row, values_only=True):
            v = row[req_col - 1] if req_col - 1 < len(row) else None
            if isinstance(v, str):
                code = v.strip()
                if code and code not in result:
                    result[code] = group
    return result


# ── 2) 네트워크 폴더 스캔 → req_code별 최신 작업지시서 파일 ──
def build_file_index(root, depth):
    files = []  # (path, mtime)
    def walk(folder, d):
        try:
            for entry in os.scandir(folder):
                if entry.is_file():
                    nm = entry.name.lower()
                    if "#" in entry.name and nm.endswith((".xlsm", ".xlsx", ".xls")) and not entry.name.startswith("~$"):
                        files.append((entry.path, entry.stat().st_mtime))
                elif entry.is_dir() and d > 1:
                    walk(entry.path, d - 1)
        except (PermissionError, OSError):
            pass
    walk(root, depth)
    return files


def find_work_order(file_index, req_code):
    low = req_code.lower()
    best, best_m = None, -1
    for path, mtime in file_index:
        if low in os.path.basename(path).lower():
            if mtime > best_m:
                best, best_m = path, mtime
    return best


# ── 3) 작업지시서 → company/site/pack range/d 배열 ───────────
def extract_job_order(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = None
    for s in wb.sheetnames:
        if s.strip() == WO_SHEET or (WO_SHEET in s):
            ws = wb[s]; break
    if ws is None:
        return None
    # C열 우선, 없으면 D열 (탈형 작업지시서는 D열)
    company = str(ws.cell(2, 3).value or ws.cell(2, 4).value or "").strip()
    site    = str(ws.cell(3, 3).value or ws.cell(3, 4).value or "").strip()
    req     = str(ws.cell(4, 3).value or ws.cell(4, 4).value or "").strip()
    if not company or not site:
        return None

    # 포장방법: 일반 데크=D5(col4), 탈형=G5(col7)
    pm_raw = str(ws.cell(5, 4).value or "").strip()
    if not pm_raw:
        pm_raw = str(ws.cell(5, 7).value or "").strip()
    pm = "슬리퍼" if "슬리퍼" in pm_raw else ("목재" if "목재" in pm_raw else "")

    rows = list(ws.iter_rows(min_row=ROW_START, max_row=ws.max_row or ROW_START, values_only=True))
    pack_sa = {}  # pack_no -> [sheets, area]
    pmin, pmax = None, None
    # B열에 유효한 패킹번호가 하나라도 있으면 B열 기준, 없으면 C열 기준 (탈형)
    def _is_pack(v):
        return isinstance(v, (int, float)) and v > 0 and int(v) == v
    use_col = COL_PACK - 1  # B열 기본
    is_tg = not any(_is_pack(row[COL_PACK - 1] if COL_PACK - 1 < len(row) else None) for row in rows)
    if is_tg:
        use_col = COL_PACK  # C열 (탈형): 패킹번호가 C열이라 장수/면적도 1칸 뒤
    col_sheets = (COL_SHEETS if not is_tg else COL_SHEETS + 1) - 1  # 0-indexed
    col_area   = (COL_AREA   if not is_tg else COL_AREA   + 1) - 1  # 0-indexed

    for row in rows:
        b = row[use_col] if use_col < len(row) else None
        if not _is_pack(b):
            continue
        p = int(b)
        n = row[col_sheets] if col_sheets < len(row) else None
        o = row[col_area]   if col_area   < len(row) else None
        s_val = int(n) if isinstance(n, (int, float)) and n > 0 else 0
        a_val = round(float(o), 2) if isinstance(o, (int, float)) and o > 0 else 0
        pack_sa[p] = [s_val, a_val]
        pmin = p if pmin is None else min(pmin, p)
        pmax = p if pmax is None else max(pmax, p)

    if pmin is None:
        return None

    d = [pack_sa.get(p, [0, 0]) for p in range(pmin, pmax + 1)]
    total_sheets = sum(x[0] for x in d)
    total_area   = round(sum(x[1] for x in d), 2)
    return {
        "type": "job_order", "company": company, "site": site, "req_code": req,
        "pack_from": pmin, "pack_to": pmax, "d": d, "pm": pm,
        "total_sheets": total_sheets, "total_area": total_area,
    }


# ── 4) Firestore 업로드 ──────────────────────────────────────
def upload(records):
    import firebase_admin
    from firebase_admin import credentials, firestore
    if not os.path.exists(SERVICE_ACCOUNT_KEY):
        print("[오류] 서비스 계정 키 파일이 없습니다:", SERVICE_ACCOUNT_KEY)
        print("  Firebase 콘솔 → 프로젝트 설정 → 서비스 계정 → 새 비공개 키 생성")
        sys.exit(1)
    cred = credentials.Certificate(SERVICE_ACCOUNT_KEY)
    if not firebase_admin._apps:
        firebase_admin.initialize_app(cred)
    db = firestore.client()
    now = datetime.datetime.now(datetime.timezone.utc)
    batch = db.batch()
    n = 0
    for rec in records:
        # 문서ID = 요청코드 기준 (같은 현장 복수 작업지시서 각각 보존)
        doc_id = make_doc_id(rec["req_code"], rec["site"])
        # d 배열 → Firestore 호환(중첩 배열 불가) → [{s,a}] 형태로 저장
        rec_fs = dict(rec)
        rec_fs["d"] = [{"s": s, "a": a} for s, a in rec["d"]]
        rec_fs["updated_at"] = now
        rec_fs["source"] = "plan_upload"
        batch.set(db.collection("job_orders").document(doc_id), rec_fs)
        n += 1
        if n % 20 == 0:
            batch.commit(); batch = db.batch()
    batch.commit()
    return n


def main():
    args = sys.argv[1:]
    # 파일 1개 추출 테스트
    if "--file" in args:
        path = args[args.index("--file") + 1]
        jo = extract_job_order(path)
        print("추출 결과:", jo if not jo else {k: (v if k != "d" else f"[{len(v)}개 패킹]") for k, v in jo.items()})
        if jo:
            print("  d(앞 8개):", jo["d"][:8])
        return

    dry = "--dry-run" in args
    clear = "--clear" in args

    # 기존 job_orders 전체 삭제
    if clear and not dry:
        import firebase_admin
        from firebase_admin import credentials, firestore as _fs
        if not firebase_admin._apps:
            firebase_admin.initialize_app(credentials.Certificate(SERVICE_ACCOUNT_KEY))
        db2 = _fs.client()
        docs = list(db2.collection("job_orders").stream())
        print(f"[삭제] 기존 job_orders {len(docs)}건 삭제 중...")
        batch2 = db2.batch()
        for i, d in enumerate(docs):
            batch2.delete(d.reference)
            if (i+1) % 400 == 0: batch2.commit(); batch2 = db2.batch()
        batch2.commit()
        print("  → 삭제 완료")

    print("[1/3] 생산계획표 요청코드 수집:", PLAN_FILE)
    reqs = collect_req_codes(PLAN_FILE)   # {req_code: 호기그룹}
    print("   →", len(reqs), "개 요청코드")

    print("[2/3] 작업지시서 폴더 스캔:", WORK_ORDER_DIRS)
    idx = []
    for d in WORK_ORDER_DIRS:
        idx += build_file_index(d, SEARCH_DEPTH)
    print("   →", len(idx), "개 파일(# 포함)")

    print("[3/3] 작업지시서 열어 면적·포장방법 추출")
    records, missing = [], []
    for code, group in reqs.items():
        path = find_work_order(idx, code)
        if not path:
            missing.append(code); continue
        try:
            jo = extract_job_order(path)
            if jo:
                jo["machine_group"] = group
                records.append(jo)
            else:
                missing.append(code + " (시트/데이터 없음)")
        except Exception as e:
            missing.append(code + " (오류: %s)" % e)

    print("\n── 추출 완료 ──")
    print("  성공:", len(records), "건 / 파일없음·실패:", len(missing), "건")
    for r in records[:60]:
        print("   - [%s] %s (%s)  pack %d~%d  %d ea / %.2f m2  pm:%s" %
              (r.get("machine_group", "?"), r["site"], r["company"], r["pack_from"], r["pack_to"],
               r["total_sheets"], r["total_area"], r.get("pm") or "-"))
    if missing:
        print("  [건너뜀]", ", ".join(missing[:30]))

    if dry:
        print("\n(--dry-run: 업로드 안 함)")
        return
    if not records:
        msg = f"[작업지시서 업로드] 업로드할 데이터 없음\n생산계획표: {os.path.basename(PLAN_FILE)}"
        print(msg); tg(msg); return
    print("\n업로드 중…")
    n = upload(records)
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    msg = (
        f"[작업지시서 업로드] 완료 ({now_str})\n"
        f"생산계획표: {os.path.basename(PLAN_FILE)}\n"
        f"업로드: {n}건"
        + (f" / 누락: {len(missing)}건" if missing else "")
    )
    if missing:
        msg += "\n누락: " + ", ".join(missing[:10]) + ("…" if len(missing) > 10 else "")
    print("job_orders upload done:", n)
    tg(msg)


if __name__ == "__main__":
    main()
